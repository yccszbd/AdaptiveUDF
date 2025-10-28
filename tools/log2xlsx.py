#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import re
import sys
import argparse
from pathlib import Path
from typing import Dict, Iterable, List
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import numbers

# -----------------------------
# 配置区：字段与解析规则
# -----------------------------
FIELDS = [
    "dataset",
    "dataname",
    "time_sum",
    "normals_correctness",
    "chamferL1_mean",
    "chamferL1_median",
    "chamferL2_mean",
    "chamferL2_median",
    "F_score_0.01",
    "F_score_0.005",
]

NUMERIC_FIELDS = {
    "time_sum",
    "normals_correctness",
    "chamferL1_mean",
    "chamferL1_median",
    "chamferL2_mean",
    "chamferL2_median",
    "F_score_0.01",
    "F_score_0.005",
}

PERCENT_FIELDS = {
    "normals_correctness",
    "F_score_0.01",
    "F_score_0.005",
}

# 默认用这个键出现时作为一个“区块”的开始（日志里普遍有它）
BLOCK_START_KEYS = {"pred_file"}

# INFO - key: value 的抓取（更宽松地允许空格）
KV_RE = re.compile(r"INFO\s*-\s*([A-Za-z0-9_.]+)\s*:\s*([^\s].*?)\s*$")
# 抓数字用，支持科学计数法；先把千位逗号去掉
NUM_RE = re.compile(r"[-+]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][-+]?\d+)?")


# -----------------------------
# 解析辅助
# -----------------------------
def parse_numeric_like(s: str) -> float:
    """
    尽可能从字符串里抽出一个数字（支持科学计数法）。
    例如：'92.5%', '92.%', '9.2e-1', '  1,234.5;' 都能取到数字。
    """
    s_clean = s.replace(",", "")  # 去掉千位分隔符
    m = NUM_RE.search(s_clean)
    if not m:
        raise ValueError(f"no numeric in '{s}'")
    return float(m.group(0))


def parse_time_seconds(s: str) -> float:
    """
    把 time_sum 解析为秒：
    - 'HH:MM:SS'、'MM:SS'、'SS'
    - '1h23m45s'、'12m34s'、'45s'
    - 纯数字（按秒）
    """
    s = s.strip().lower()
    # 冒号形式
    if ":" in s:
        parts = [p for p in s.split(":") if p != ""]
        if not all(p.replace(".", "", 1).isdigit() for p in parts):
            # 混有非数字，回退
            return parse_numeric_like(s)
        parts = list(map(float, parts))
        if len(parts) == 3:
            h, m, sec = parts
            return h * 3600 + m * 60 + sec
        elif len(parts) == 2:
            m, sec = parts
            return m * 60 + sec
        elif len(parts) == 1:
            return parts[0]
        else:
            return parse_numeric_like(s)

    # XhYmZs 形式
    h = m = sec = 0.0
    has_flag = False
    mh = re.search(r"([-+]?\d+(?:\.\d*)?)h", s)
    if mh:
        h = float(mh.group(1))
        has_flag = True
    mm = re.search(r"([-+]?\d+(?:\.\d*)?)m", s)
    if mm:
        m = float(mm.group(1))
        has_flag = True
    ms = re.search(r"([-+]?\d+(?:\.\d*)?)s", s)
    if ms:
        sec = float(ms.group(1))
        has_flag = True
    if has_flag:
        return h * 3600 + m * 60 + sec

    # 纯数字（按秒）
    return parse_numeric_like(s)


def coerce_value(field: str, raw: str):
    """
    把原始字符串转换为适合写 Excel 的数值或文本。
    - 百分比字段：统一转成 0~1 的浮点，并设置百分比格式由外层完成
    - time_sum：解析成秒（浮点）
    - 其他数值字段：解析成 float
    - 解析失败则返回原字符串（文本）
    """
    txt = raw.strip().rstrip(",;")  # 去掉常见尾巴
    if field not in NUMERIC_FIELDS:
        return txt, None  # (值, number_format)

    try:
        if field in PERCENT_FIELDS:
            # 允许 92% / 92 / 0.92 / 9.2e-1 等
            if "%" in txt:
                v = parse_numeric_like(txt) / 100.0
            else:
                v = parse_numeric_like(txt)
                if v > 1.0:
                    v = v / 100.0
            return v, numbers.FORMAT_PERCENTAGE_00

        if field == "time_sum":
            v = parse_time_seconds(txt)
            # 秒用常规数值格式；如需 mm:ss 可额外自定义
            return v, numbers.FORMAT_NUMBER_00

        # 其他普通数值
        v = parse_numeric_like(txt)
        return v, None

    except Exception:
        # 解析失败，回落为文本
        return txt, None


# -----------------------------
# 日志 -> 记录（流式）
# -----------------------------
def iter_records(fp, start_keys: set) -> Iterable[Dict[str, str]]:
    """
    逐行扫描：遇到 start_key（默认 pred_file）认为是新区块起始；
    把区块内出现的 INFO - key: value 聚合为一条记录。
    """
    current: Dict[str, str] = {}
    for line in fp:
        m = KV_RE.search(line)
        if not m:
            continue
        key, val = m.groups()

        # 如果遇到区块起始键且当前已有内容，则先产出上一条
        if key in start_keys and current:
            yield current
            current = {}

        # 收集字段（不用的键可以忽略；也可把 start_key 值带上）
        if key in FIELDS or key in start_keys:
            current[key] = val

    if current:
        yield current


# -----------------------------
# 写 Excel
# -----------------------------
def write_to_xlsx(records: List[Dict[str, str]], out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "results"

    # 表头
    bold = Font(bold=True)
    for col, field in enumerate(FIELDS, start=1):
        c = ws.cell(row=1, column=col, value=field)
        c.font = bold

    # 数据
    for row_idx, rec in enumerate(records, start=2):
        for col_idx, field in enumerate(FIELDS, start=1):
            raw = rec.get(field, "")
            value, numfmt = coerce_value(field, raw) if raw != "" else ("", None)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if numfmt:
                cell.number_format = numfmt

    # 冻结首行+自动筛选
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(FIELDS))}{len(records) + 1}"

    # 自动列宽
    for col_idx, field in enumerate(FIELDS, start=1):
        max_len = len(field)
        for row in range(2, len(records) + 2):
            v = ws.cell(row=row, column=col_idx).value
            s = f"{v}" if v is not None else ""
            if ws.cell(row=row, column=col_idx).number_format in (
                numbers.FORMAT_PERCENTAGE_00,
                numbers.FORMAT_NUMBER_00,
            ):
                # 数值列留点余量
                s = s if not isinstance(v, float) else f"{v:.4g}"
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    wb.save(out_path)


# -----------------------------
# 主流程
# -----------------------------
def main():
    ap = argparse.ArgumentParser(description="Parse training logs to XLSX.")
    ap.add_argument("log", nargs="?", default="training.log", help="日志文件路径")
    ap.add_argument(
        "-o", "--out", default=None, help="输出 XLSX 路径（默认与日志同名）"
    )
    ap.add_argument(
        "--start-key", default="pred_file", help="作为区块起始的键名（默认：pred_file）"
    )
    ap.add_argument("--encoding", default="utf-8", help="读取编码（默认 utf-8）")
    args = ap.parse_args()

    log_path = Path(args.log)
    if not log_path.exists():
        print(f"❌ 找不到日志文件：{log_path}")
        sys.exit(1)

    out_xlsx = (
        Path(args.out)
        if args.out
        else log_path.with_name(f"{log_path.stem}_result.xlsx")
    )
    start_keys = {args.start_key}

    # 流式读取并聚合记录（只保留 FIELDS 中定义的列）
    records: List[Dict[str, str]] = []
    with log_path.open("r", encoding=args.encoding, errors="ignore") as fp:
        for rec in iter_records(fp, start_keys=start_keys):
            # 只保留我们需要的字段
            slim = {k: v for k, v in rec.items() if k in FIELDS}
            if slim:
                records.append(slim)

    if not records:
        print("⚠️ 日志中没有匹配到任何数据（检查 INFO - key: value 的格式或 start-key）")
        sys.exit(0)

    write_to_xlsx(records, out_xlsx)
    print(f"✅ 已写入 {out_xlsx.resolve()}，可直接用 Excel 打开查看")


if __name__ == "__main__":
    main()
